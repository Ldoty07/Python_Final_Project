# Import Workbook
from openpyxl import Workbook

def full_inventory_report(product_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Full Inventory Report"

    sheet.append(["Full Inventory Report"])
    sheet.append([])
    sheet.append(["Product ID", "Product Name", "Category", "Quantity In Stock", "Unit Price"])

    total_inventory_value = 0

    for product in product_list:
        sheet.append([product.product_id, product.product_name, product.category, product.qty_in_stock, product.unit_price])
        total_inventory_value += product.qty_in_stock * product.unit_price

    sheet.append(["Total Inventory Value", total_inventory_value])

    workbook.save("full_inventory_report.xlsx")
    print("Data exported to full_inventory_report.xlsx")

def low_stock_report(product_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Low Stock Report"

    sheet.append(["Low Stock Report"])
    sheet.append([])
    sheet.append(["Product ID", "Product Name", "Category", "Quantity In Stock", "Reorder Level", "Reorder Quantity"])

    for product in product_list:
        sheet.append([product.product_id, product.product_name, product.category, product.qty_in_stock, product.reorder_lvl, product.reorder_qty])

    workbook.save("low_stock_report.xlsx")
    print("Data exported to low_stock_report.xlsx")

def open_PO_report(pending_PO_list, PO_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Open Purchase Order Report"

    sheet.append(["Open Purchase Order Report"])
    sheet.append([])
    sheet.append(["PO number", "Vendor ID", "Total Cost", "Status"])

    open_PO_list = []

    for pending_PO in pending_PO_list:
        open_PO_list.append(pending_PO)

    for sent_PO in PO_list:
        if sent_PO.status == "Sent":
            open_PO_list.append(sent_PO)

    for open_PO in open_PO_list:
        sheet.append([open_PO.PO_number, open_PO.vendor_id, open_PO.total_cost, open_PO.status])

    workbook.save("open_po_report.xlsx")
    print("Data exported to open_po_report.xlsx")